import os
from openpyxl import load_workbook
from Statistics.variance import variance
from Statistics.standard_deviation import standard_deviation, standard_deviation_from_known_variance

if __name__ == '__main__':
    path_to_excel_file = __file__.replace('\\' + __file__.split('\\')[-1], '\\xlsx\\學校防疫政策滿意度調查_term_report.xlsx')
    path_to_excel_file = os.path.abspath(path_to_excel_file)
    #print(path_to_excel_file) # DEBUG
    range_start = 3
    range_end   = 13
    
    try:
        # southern regions 
        range_cells_s = [] 
        range_cells_s += ['E'+str(num) for num in range(range_start, range_end)] # initialize a list with 'list comprehension' in Python
        range_cells_s += ['H'+str(num) for num in range(range_start, range_end)] # initialize a list with 'list comprehension' in Python
        range_cells_s += ['K'+str(num) for num in range(range_start, range_end)] # initialize a list with 'list comprehension' in Python
        range_cells_s += ['L'+str(num) for num in range(range_start, range_end)] # initialize a list with 'list comprehension' in Python
        range_cells_s += ['O'+str(num) for num in range(range_start, range_end)] # initialize a list with 'list comprehension' in Python
        #print(range_cells_s) # DEBUG
        #for coord in range_cells_s:
        #    print(coord) # DEBUG
        
        wb = load_workbook(path_to_excel_file, data_only=True)
        ws = wb['樞紐分析']
        # variance 
        var_s = variance(worksheet=ws, aver_cell='V14', range_cells=range_cells_s)
        # standard deviation 
        sdev_s = standard_deviation_from_known_variance(variance=var_s)
        
        ws['V16'].value = var_s
        ws['V15'].value = sdev_s

        print("(此抽樣調查中) 南部 地區的學生滿意度變異數 = " + str(var_s))
        print("(此抽樣調查中) 南部 地區的學生滿意度標準差 = " + str(sdev_s))

        # other regions 
        range_cells_other_region = []
        range_cells_other_region += ['B'+str(num) for num in range(range_start, range_end)] # initialize a list with 'list comprehension' in Python
        range_cells_other_region += ['C'+str(num) for num in range(range_start, range_end)] # initialize a list with 'list comprehension' in Python
        range_cells_other_region += ['D'+str(num) for num in range(range_start, range_end)] # initialize a list with 'list comprehension' in Python
        range_cells_other_region += ['F'+str(num) for num in range(range_start, range_end)] # initialize a list with 'list comprehension' in Python
        range_cells_other_region += ['G'+str(num) for num in range(range_start, range_end)] # initialize a list with 'list comprehension' in Python
        range_cells_other_region += ['I'+str(num) for num in range(range_start, range_end)] # initialize a list with 'list comprehension' in Python
        range_cells_other_region += ['J'+str(num) for num in range(range_start, range_end)] # initialize a list with 'list comprehension' in Python
        range_cells_other_region += ['M'+str(num) for num in range(range_start, range_end)] # initialize a list with 'list comprehension' in Python
        range_cells_other_region += ['N'+str(num) for num in range(range_start, range_end)] # initialize a list with 'list comprehension' in Python
        range_cells_other_region += ['P'+str(num) for num in range(range_start, range_end)] # initialize a list with 'list comprehension' in Python
        #print(range_cells_other_region) # DEBUG
        #for coord in range_cells_other_region:
        #    print(coord) # DEBUG
        
        # variance 
        var_other_region = variance(worksheet=ws, aver_cell='W14', range_cells=range_cells_other_region)
        # standard deviation 
        sdev_other_region = standard_deviation_from_known_variance(variance=var_other_region)
        
        ws['W16'].value = var_other_region
        ws['W15'].value = sdev_other_region
        
        print("(此抽樣調查中) 其他 地區的學生滿意度變異數 = " + str(var_other_region))
        print("(此抽樣調查中) 其他 地區的學生滿意度標準差 = " + str(sdev_other_region))

        # all regions 
        range_cells = [] # all regions 
        range_cells += range_cells_s
        range_cells += range_cells_other_region
        #print(range_cells) # DEBUG
        #for coord in range_cells:
        #    print(coord) # DEBUG
        
        # variance 
        var = variance(worksheet=ws, aver_cell='X14', range_cells=range_cells)
        # standard deviation 
        sdev = standard_deviation_from_known_variance(variance=var)
        
        ws['X16'].value = var
        ws['X15'].value = sdev
        
        print("(此抽樣調查中) 全部 地區的學生滿意度變異數 = " + str(var))
        print("(此抽樣調查中) 全部 地區的學生滿意度標準差 = " + str(sdev))

        # save to .xlsx file 
        wb.save(path_to_excel_file)
        # 告知使用者本程式正常執行完畢 
        print("\n[Python程式完整執行完畢]\n(file path: " + path_to_excel_file + " )\n")
    except Exception as ex:
        print("\nPython程式發生錯誤: \n" + str(ex) + "\n")
